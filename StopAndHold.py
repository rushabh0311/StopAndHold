import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook

file_path = input("Enter the Excel file: ")
save_file_path = input("Enter the output Excel file name: ")

wb = Workbook()
ws = wb.active
ws.append(['Sheet Name', 'Time', 'Inclination Angle', 'Target Angle', 'Force', 'Original Force'])

def detect_and_plot_84_peak(df, ax):
    df['Original_Angle1'] = pd.to_numeric(df['Angle1'], errors='coerce')
    df['Smoothed'] = pd.to_numeric(df['Smoothed'], errors='coerce')
    df['Time'] = pd.to_numeric(df['Time'], errors='coerce')

    target_df = df[(df['Smoothed'] >= 82.5) & (df['Smoothed'] <= 84)].copy()
    if target_df.empty:
        print("No 84° peak found in smoothed range.")
        return None

    closest_idx = (target_df['Smoothed'] - 84).abs().idxmin()
    closest_row = target_df.loc[closest_idx]

    return {
        'Time': closest_row['Time'],
        'Inclination Angle': 84,
        'Target Angle': closest_row['Smoothed'],
        'Force': abs(closest_row['Original_Angle1']),
        'Original Force': closest_row['Original_Angle1']
    }

def process_sheet(sheet_name, df, inclination_angles):
    df.columns = df.columns.str.strip()

    df['Original_Angle1'] = pd.to_numeric(df['Angle1'], errors='coerce')
    df['Original_Angle'] = pd.to_numeric(df['Angle'], errors='coerce')
    df['Smoothed'] = pd.to_numeric(df['Smoothed'], errors='coerce')
    df['Time'] = pd.to_numeric(df['Time'], errors='coerce')

    df['Angle1'] = df['Original_Angle1'].abs()
    df['Angle'] = df['Original_Angle'].abs()

    all_results = []

    def find_cycle_peaks(input_angle):
        if input_angle == 84:
            target_df = df[(df['Smoothed'] >= 82.5) & (df['Smoothed'] <= 84)].copy()
            if target_df.empty:
                print("No data found in the 82.5 - 84 smoothed range.")
                return None
            closest_idx = (target_df['Smoothed'] - input_angle).abs().idxmin()
            closest_row = target_df.loc[closest_idx]
            exact_target = closest_row['Smoothed']
            print(f"Closest smoothed value to 84 in range (82.5–84): {exact_target:.3f}")
        else:
            if sheet_name.lower().strip() == 'sheet3':
                dynamic_tolerance = 0.4
            else:
                smoothed_range = df['Smoothed'].max() - df['Smoothed'].min()
                dynamic_tolerance = max(0.2, min(1.2, smoothed_range * 0.01))

            print(f"[{sheet_name}] Dynamic tolerance used: ±{dynamic_tolerance:.3f}")
            df['Difference'] = abs(df['Smoothed'] - input_angle)
            target_df = df[(df['Smoothed'] >= input_angle - dynamic_tolerance) &
                           (df['Smoothed'] <= input_angle + dynamic_tolerance)].copy()

            if target_df.empty:
                print(f"No data found within ±{dynamic_tolerance:.3f} range.")
                return None
            exact_target = input_angle

        # Positive cycle
        positive_df = target_df[target_df['Original_Angle1'] > 0]
        positive_peak = None
        if not positive_df.empty:
            positive_peak = positive_df.loc[positive_df['Original_Angle1'].idxmax()]

        # Negative cycle
        negative_df = target_df[target_df['Original_Angle1'] < 0]
        negative_peak = None
        if not negative_df.empty:
            negative_peak = negative_df.loc[negative_df['Original_Angle1'].idxmin()]

        if positive_peak is not None and negative_peak is not None:
            if abs(positive_peak['Time'] - negative_peak['Time']) < 1:
                print("Skipping points - Time gap < 1 second.")
                return None

            return [
                {
                    'Time': positive_peak['Time'],
                    'Inclination Angle': input_angle,
                    'Target Angle': positive_peak['Smoothed'],
                    'Force': positive_peak['Angle1'],
                    'Original Force': positive_peak['Original_Angle1']
                },
                {
                    'Time': negative_peak['Time'],
                    'Inclination Angle': input_angle,
                    'Target Angle': negative_peak['Smoothed'],
                    'Force': negative_peak['Angle1'],
                    'Original Force': negative_peak['Original_Angle1']
                }
            ]
        else:
            print("Could not find both positive and negative peaks.")
            return None

    for input_angle in inclination_angles:
        print(f"\nProcessing angle {input_angle} for sheet {sheet_name}")
        result = find_cycle_peaks(input_angle)
        if result:
            all_results.extend(result)
            print(pd.DataFrame(all_results))
        else:
            print(f"No suitable peak values found for angle {input_angle}.")

    if not all_results:
        return None, []

    # Ask about threshold lines
    add_threshold_lines = False
    user_input = input("Do you want to add horizontal dashed lines at +5, -5, +15, and -15? (yes/no): ").strip().lower()
    if user_input == 'yes':
        add_threshold_lines = True
        var1 = int(input("Enter the first threshold value for line 1: "))
        var2 = int(input("Enter the second threshold value for line 2: "))
        var3 = int(input("Enter the third threshold value for line 3: "))
        var4 = int(input("Enter the fourth threshold value for line 4: "))
    else:
        print("Skipping threshold lines.")

    # Plotting
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(df['Smoothed'], df['Original_Angle1'], color='lightblue', linestyle='-', alpha=0.7)

    highlighted_points = pd.DataFrame(all_results)
    for i in range(0, len(highlighted_points), 2):
        red_point = highlighted_points.iloc[i]
        green_point = highlighted_points.iloc[i + 1]

        ax.scatter(red_point['Target Angle'], red_point['Original Force'], color='red', label='' if i == 0 else "", zorder=5)
        ax.scatter(green_point['Target Angle'], green_point['Original Force'], color='green', label='' if i == 0 else "", zorder=5)
        ax.plot([red_point['Target Angle'], green_point['Target Angle']],
                [red_point['Original Force'], green_point['Original Force']],
                color='gray', linestyle='--', linewidth=1, zorder=4)


    # Draw threshold lines if selected
    if add_threshold_lines:
        thresholds = [(var1, 'orange'), (var2, 'blue'), (var3, 'purple'), (var4, 'brown')]
        for y, color in thresholds:
            ax.axhline(y=y, color=color, linestyle='--', linewidth=1.5, label=f'y = {y}')

    min_force = min(df['Original_Angle1'].min(), highlighted_points['Original Force'].min()) - 5
    max_force = max(df['Original_Angle1'].max(), highlighted_points['Original Force'].max()) + 5
    ax.set_ylim(min_force, max_force)

    ax.set_xlabel('Smoothed')
    ax.set_ylabel('Angle1')
    ax.set_title(f'Graph for {sheet_name}: Smoothed vs Angle1')
    ax.grid(True)
    ax.legend()
    fig.tight_layout()
    plt.savefig(f'graph_{sheet_name}.png')
    plt.show()

    for row in all_results:
        ws.append([sheet_name, row['Time'], row['Inclination Angle'], row['Target Angle'], row['Force'], row['Original Force']])

    return sheet_name, all_results

# Input inclination angles
while True:
    try:
        angles_input = input("Enter 4 inclination angles separated by commas (e.g., 30,45,60,75): ")
        inclination_angles = [float(a.strip()) for a in angles_input.split(',')]
        if len(inclination_angles) != 4:
            print("❗ Please enter exactly 4 angles.")
            continue
        break
    except ValueError:
        print("❗ Invalid input. Please enter numeric values only.")

# Process all sheets
book = load_workbook(file_path)
sheet_names = book.sheetnames
for i, sheet_name in enumerate(sheet_names[1:], start=1):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    result, sheet_results = process_sheet(sheet_name, df, inclination_angles)
    if result is None:
        break

wb.save(save_file_path)
print("✅ Processing complete and data saved.")